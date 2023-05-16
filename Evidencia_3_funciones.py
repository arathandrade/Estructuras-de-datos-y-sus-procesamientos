#ARATH EMMANUEL SANCHEZ ANDRADE 2025792 GRUPO 32
import csv
import openpyxl
# VARIABLES
lista = []
global id 

autores=[]
global id_autor 

generos=[]
global id_genero

#FUNCIONES

def registrar_ejemplar():
    print(id)
    titulo_t = input("titulo del libro: ")
    autor_t = input("autor del libro: ")
    genero_t = input("genero del libro: ")
    año_de_publicacion_t = input("año de publicacion del libro: ")
    isbn_t = input("ISBN del libro: ")
    fecha_de_adquisicion_t = input("fecha de adquisicion del libro: ")

    libro = dict(id= nuevo_id(), titulo = titulo_t.upper(), autor = autor_t.upper(), genero = genero_t.upper(), año_de_publicacion = año_de_publicacion_t.upper(), isbn = isbn_t.upper(), fecha_de_adquisicion = fecha_de_adquisicion_t.upper())
    
    lista.append(libro)

def registrar_autor():

    nombre_t = input("Nombre del autor: ")
    apellido_t = input("Apellido del autor: ")

    autor= dict(id=nuevo_id_autor(), nombre = nombre_t.upper(), apellido = apellido_t.upper())

    autores.append(autor)

def registrar_genero():

    genero_t = input("Genero del libro: ")

    genero = dict(id=nuevo_id_genero(), genero = genero_t.upper())

    generos.append(genero)

def catalogo_completo():
    print("ID\tTITULO\tAUTOR\tGENERO\tAÑO DE PUBLICACION\tISBN\tFECHA DE ADQUISICION")

    for libro in lista:
        print(libro["id"],"\t",libro["titulo"],"\t",libro["autor"],"\t",libro["genero"],"\t",libro["año_de_publicacion"],"\t",libro["isbn"],"\t",libro["fecha_de_adquisicion"])
    print()
    while True:
        print("Accion sobre el reporte:")
        print("1. Exportar reporte a formato CSV")
        print("2. Exportar reporte a formato MsExcel")
        print("0. No exportar reporte")
                    

        opcion = input("Elige una opción: ")

        if opcion == "1":
            print("Has elegido la opción 1")
            catalogo_completo_csv() #CATALOGO COMPLETO 
        elif opcion == "2":
            print("Has elegido la opción 2")
            catalogo_completo_excel()
        elif opcion == "0":
            print("Regresando al menú anterior...")
            break
        else:
            print("Opción no válida. Por favor, elige de nuevo.")

def catalogo_completo_csv():
    # Abre el archivo CSV en modo de escritura
    with open('catalogo_completo.csv', 'w', newline='') as archivo_csv: 
    
    # Crea un escritor CSV
        escritor_csv = csv.writer(archivo_csv, delimiter=',')

        encabezado = ['ID', 'TITULO', 'AUTOR', 'GENERO', 'AÑO DE PUBLICACION', 'ISBN', 'FECHA DE ADQUISICION']
    
        escritor_csv.writerow(encabezado)

        for libro in lista:
            renglon=[libro["id"],libro["titulo"],libro["autor"], libro["genero"],libro["año_de_publicacion"],libro["isbn"],libro["fecha_de_adquisicion"]]
            escritor_csv.writerow(renglon)
    print("catalogo completo guardado")

def catalogo_completo_excel():

# Lista de diccionarios
    datos = lista

# Crea un nuevo libro de Excel
    libro_Excel = openpyxl.Workbook()

# Selecciona la hoja activa
    hoja = libro_Excel.active

# Escribe la fila de encabezado en la hoja
    encabezado = ['ID', 'TITULO', 'AUTOR', 'GENERO', 'AÑO DE PUBLICACION', 'ISBN', 'FECHA DE ADQUISICION']
    hoja.append(encabezado)

# Escribe los datos de cada fila en la hoja
    for libro in datos:
        valores = [libro["id"],libro["titulo"],libro["autor"], libro["genero"],libro["año_de_publicacion"],libro["isbn"],libro["fecha_de_adquisicion"]]
        hoja.append(valores)

# Guarda el libro de Excel en un archivo
    libro_Excel.save('catalogo_completo_excel.xlsx')

    print("catalogo completo guardado en excel")

def reporte_autor():
    autores = set()
    for libro in lista:
        autores.add(libro["autor"])
    print("autores:")
    for autor in autores:
        print(autor)

    buscar = input("qué autor quieres buscar?")
    print("ID\tTITULO\tAUTOR\tGENERO\tAÑO DE PUBLICACION\tISBN\tFECHA DE ADQUISICION")

    for libro in lista:
        if libro["autor"] == buscar.upper():  
            print(libro["id"],"\t",libro["titulo"],"\t",libro["autor"],"\t",libro["genero"],"\t",libro["año_de_publicacion"],"\t",libro["isbn"],"\t",libro["fecha_de_adquisicion"])
    print()

    while True:
        print("Accion sobre el reporte:")
        print("1. Exportar reporte a formato CSV")
        print("2. Exportar reporte a formato MsExcel")
        print("0. No exportar reporte")
                    

        opcion = input("Elige una opción: ")

        if opcion == "1":
            print("Has elegido la opción 1")
            reporte_autor_csv(buscar) #CATALOGO COMPLETO 
        elif opcion == "2":
            print("Has elegido la opción 2")
            reporte_autor_excel(buscar)
        elif opcion == "0":
            print("Regresando al menú anterior...")
            break
        else:
            print("Opción no válida. Por favor, elige de nuevo.")

def reporte_autor_csv(buscar):
    # Abre el archivo CSV en modo de escritura
    with open('reporte_autor.csv', 'w', newline='') as archivo_csv: 
    
    # Crea un escritor CSV
        escritor_csv = csv.writer(archivo_csv, delimiter=',')

        encabezado = ['ID', 'TITULO', 'AUTOR', 'GENERO', 'AÑO DE PUBLICACION', 'ISBN', 'FECHA DE ADQUISICION']
    
        escritor_csv.writerow(encabezado)

        for libro in lista:
            if libro["autor"] == buscar.upper():  
                renglon=[libro["id"],libro["titulo"],libro["autor"], libro["genero"],libro["año_de_publicacion"],libro["isbn"],libro["fecha_de_adquisicion"]]
                escritor_csv.writerow(renglon)
    print("reporte por autor guardado en csv")

def reporte_autor_excel(buscar):

# Lista de diccionarios
    datos = lista

# Crea un nuevo libro de Excel
    libro_Excel = openpyxl.Workbook()

# Selecciona la hoja activa
    hoja = libro_Excel.active

# Escribe la fila de encabezado en la hoja
    encabezado = ['ID', 'TITULO', 'AUTOR', 'GENERO', 'AÑO DE PUBLICACION', 'ISBN', 'FECHA DE ADQUISICION']
    hoja.append(encabezado)

# Escribe los datos de cada fila en la hoja
    for libro in datos:
        if libro["autor"] == buscar.upper():  
            valores = [libro["id"],libro["titulo"],libro["autor"], libro["genero"],libro["año_de_publicacion"],libro["isbn"],libro["fecha_de_adquisicion"]]
            hoja.append(valores)

# Guarda el libro de Excel en un archivo
    libro_Excel.save('reporte_autor_excel.xlsx')

    print("reporte por autor guardado en excel")

def reporte_genero():
    generos = set()
    for libro in lista:
        generos.add(libro["genero"])
    print("generos:")
    for genero in generos:
        print(genero)

    buscar = input("qué genero quieres buscar?")
    print("ID\tTITULO\tAUTOR\tGENERO\tAÑO DE PUBLICACION\tISBN\tFECHA DE ADQUISICION")

    for libro in lista:
        if libro["genero"] == buscar.upper():  
            print(libro["id"],"\t",libro["titulo"],"\t",libro["autor"],"\t",libro["genero"],"\t",libro["año_de_publicacion"],"\t",libro["isbn"],"\t",libro["fecha_de_adquisicion"])
    print()

    while True:
        print("Accion sobre el reporte:")
        print("1. Exportar reporte a formato CSV")
        print("2. Exportar reporte a formato MsExcel")
        print("0. No exportar reporte")
                    

        opcion = input("Elige una opción: ")

        if opcion == "1":
            print("Has elegido la opción 1")
            reporte_genero_csv(buscar) 
        elif opcion == "2":
            print("Has elegido la opción 2")
            reporte_genero_excel(buscar)
        elif opcion == "0":
            print("Regresando al menú anterior...")
            break
        else:
            print("Opción no válida. Por favor, elige de nuevo.")

def reporte_genero_csv(buscar):
    # Abre el archivo CSV en modo de escritura
    with open('reporte_genero.csv', 'w', newline='') as archivo_csv: 
    
    # Crea un escritor CSV
        escritor_csv = csv.writer(archivo_csv, delimiter=',')

        encabezado = ['ID', 'TITULO', 'AUTOR', 'GENERO', 'AÑO DE PUBLICACION', 'ISBN', 'FECHA DE ADQUISICION']
    
        escritor_csv.writerow(encabezado)

        for libro in lista:
            if libro["genero"] == buscar.upper():  
                renglon=[libro["id"],libro["titulo"],libro["autor"], libro["genero"],libro["año_de_publicacion"],libro["isbn"],libro["fecha_de_adquisicion"]]
                escritor_csv.writerow(renglon)
    print("reporte por genero guardado en csv")

def reporte_genero_excel(buscar):

# Lista de diccionarios
    datos = lista

# Crea un nuevo libro de Excel
    libro_Excel = openpyxl.Workbook()

# Selecciona la hoja activa
    hoja = libro_Excel.active

# Escribe la fila de encabezado en la hoja
    encabezado = ['ID', 'TITULO', 'AUTOR', 'GENERO', 'AÑO DE PUBLICACION', 'ISBN', 'FECHA DE ADQUISICION']
    hoja.append(encabezado)

# Escribe los datos de cada fila en la hoja
    for libro in datos:
        if libro["genero"] == buscar.upper():  
            valores = [libro["id"],libro["titulo"],libro["autor"], libro["genero"],libro["año_de_publicacion"],libro["isbn"],libro["fecha_de_adquisicion"]]
            hoja.append(valores)

# Guarda el libro de Excel en un archivo
    libro_Excel.save('reporte_genero_excel.xlsx')

    print("reporte por genero guardado en excel")


def reporte_año():
    buscar = input("qué año quieres buscar?")
    print("ID\tTITULO\tAUTOR\tGENERO\tAÑO DE PUBLICACION\tISBN\tFECHA DE ADQUISICION")

    for libro in lista:
        if libro["año_de_publicacion"] == buscar.upper():  
            print(libro["id"],"\t",libro["titulo"],"\t",libro["autor"],"\t",libro["genero"],"\t",libro["año_de_publicacion"],"\t",libro["isbn"],"\t",libro["fecha_de_adquisicion"])
    print()

    while True:
        print("Accion sobre el reporte:")
        print("1. Exportar reporte a formato CSV")
        print("2. Exportar reporte a formato MsExcel")
        print("0. No exportar reporte")
                    

        opcion = input("Elige una opción: ")

        if opcion == "1":
            print("Has elegido la opción 1")
            reporte_año_csv(buscar) 
        elif opcion == "2":
            print("Has elegido la opción 2")
            reporte_año_excel(buscar)
        elif opcion == "0":
            print("Regresando al menú anterior...")
            break
        else:
            print("Opción no válida. Por favor, elige de nuevo.")

def reporte_año_csv(buscar):
    # Abre el archivo CSV en modo de escritura
    with open('reporte_año.csv', 'w', newline='') as archivo_csv: 
    
    # Crea un escritor CSV
        escritor_csv = csv.writer(archivo_csv, delimiter=',')

        encabezado = ['ID', 'TITULO', 'AUTOR', 'GENERO', 'AÑO DE PUBLICACION', 'ISBN', 'FECHA DE ADQUISICION']
    
        escritor_csv.writerow(encabezado)

        for libro in lista:
            if libro["año_de_publicacion"] == buscar.upper():  
                renglon=[libro["id"],libro["titulo"],libro["autor"], libro["genero"],libro["año_de_publicacion"],libro["isbn"],libro["fecha_de_adquisicion"]]
                escritor_csv.writerow(renglon)
    print("reporte por año de publicacion guardado en csv")

def reporte_año_excel(buscar):

# Lista de diccionarios
    datos = lista

# Crea un nuevo libro de Excel
    libro_Excel = openpyxl.Workbook()

# Selecciona la hoja activa
    hoja = libro_Excel.active

# Escribe la fila de encabezado en la hoja
    encabezado = ['ID', 'TITULO', 'AUTOR', 'GENERO', 'AÑO DE PUBLICACION', 'ISBN', 'FECHA DE ADQUISICION']
    hoja.append(encabezado)

# Escribe los datos de cada fila en la hoja
    for libro in datos:
        if libro["año_de_publicacion"] == buscar.upper():  
            valores = [libro["id"],libro["titulo"],libro["autor"], libro["genero"],libro["año_de_publicacion"],libro["isbn"],libro["fecha_de_adquisicion"]]
            hoja.append(valores)

# Guarda el libro de Excel en un archivo
    libro_Excel.save('reporte_año_excel.xlsx')

    print("reporte por año de publicacion guardado en excel")

def reporte_titulo():
    titulos = set()
    for libro in lista:
        titulos.add(libro["titulo"])
    print("titulos:")
    for titulo in titulos:
        print(titulo)

    buscar = input("qué titulo quieres buscar?")
    

    for libro in lista:
        if libro["titulo"] == buscar.upper():  
            print("id:\t", libro["id"])
            print("titulo:\t",libro["titulo"])
            print("autor:\t",libro["autor"])
            print("genero:\t",libro["genero"])
            print("año_de_publicacion:\t",libro["año_de_publicacion"])
            print("isbn:\t",libro["isbn"])
            print("fecha_de_adquisicion:\t",libro["fecha_de_adquisicion"])
    print()

def reporte_isbn():
    buscar = input("qué ISBN quieres buscar?")

    for libro in lista:
        if libro["isbn"] == buscar.upper():  
            print("id:\t", libro["id"])
            print("titulo:\t",libro["titulo"])
            print("autor:\t",libro["autor"])
            print("genero:\t",libro["genero"])
            print("año_de_publicacion:\t",libro["año_de_publicacion"])
            print("isbn:\t",libro["isbn"])
            print("fecha_de_adquisicion:\t",libro["fecha_de_adquisicion"]) 
    print()

def nuevo_id():
    global id

    id = id + 1
    return id

def nuevo_id_autor():
    global id_autor
    
    id_autor = id_autor + 1
    return id_autor

def nuevo_id_genero():
    global id_genero
    
    id_genero = id_genero + 1
    return id_genero

def guardar_datos():
    import csv

# Lista de diccionarios
    datos = lista

    if datos:
# Abre el archivo CSV en modo de escritura
        with open('datos_ejemplares.csv', 'w', newline='') as archivo_csv:
    
    # Crea un escritor CSV
            escritor_csv = csv.DictWriter(archivo_csv, fieldnames=['id', 'titulo', 'autor', 'genero', 'año_de_publicacion', 'isbn', 'fecha_de_adquisicion'])
    
    # Escribe la fila de encabezado
            escritor_csv.writeheader()
    
    # Escribe los datos de cada fila
            for fila in datos:
                escritor_csv.writerow(fila)
            print("datos de libros guardados")

    # Lista de diccionarios
    datos = autores

    if datos:
# Abre el archivo CSV en modo de escritura
        with open('datos_autores.csv', 'w', newline='') as archivo_csv:
    
    # Crea un escritor CSV
            escritor_csv = csv.DictWriter(archivo_csv, fieldnames=['id', 'nombre', 'apellido'])
    
    # Escribe la fila de encabezado
            escritor_csv.writeheader()
    
    # Escribe los datos de cada fila
            for fila in datos:
                escritor_csv.writerow(fila)
            print("datos de autores guardados")

    # Lista de diccionarios
    datos = generos
    if datos:
# Abre el archivo CSV en modo de escritura
        with open('datos_generos.csv', 'w', newline='') as archivo_csv:
    
    # Crea un escritor CSV
            escritor_csv = csv.DictWriter(archivo_csv, fieldnames=['id', 'genero'])
    
    # Escribe la fila de encabezado
            escritor_csv.writeheader()
    
    # Escribe los datos de cada fila
            for fila in datos:
                escritor_csv.writerow(fila)
            print("datos de generos guardados")    

def leer_datos():
    
    try:
# Abre el archivo CSV en modo de lectura
        with open('datos_ejemplares.csv', newline='') as archivo_csv:
    
    # Crea un lector CSV
            lector_csv = csv.DictReader(archivo_csv)     
    
    # Lee cada fila del archivo CSV
            for fila in lector_csv:
        
        # Agrega la fila como un diccionario a la lista de datos
                lista.append(fila)

    except FileNotFoundError:
        
        print("No hay datos previos de libros")

    try:
    
        with open('datos_autores.csv', newline='') as archivo_csv:

            lector_csv = csv.DictReader(archivo_csv)

            for fila in lector_csv:

                autores.append(fila)

    except FileNotFoundError:

        print("No hay datos previos de autores")            

    try:
    
        with open('datos_generos.csv', newline='') as archivo_csv:

            lector_csv = csv.DictReader(archivo_csv)

            for fila in lector_csv:

                generos.append(fila)

    except FileNotFoundError:

        print("No hay datos previos de generos")            

        
# Imprime los datos almacenados en la lista
    #print(lista)

    global id 
    if lista:
        id = int(lista[-1]['id'])
    else:
        id=0 
    #print(id)

    global id_autor
    if autores:
        id_autor = int(autores[-1]['id'])
    else:
        id_autor=0
    #print(id_autor)

    global id_genero
    if generos:
        id_genero = int(generos[-1]['id'])
    else:
        id_genero=0
    #print(id_genero)

